import csv
import datetime
import io
import json
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

from classifier import classify_ensemble, classify_per_chapter_ensemble, apply_hts_overrides
from config import get_api_key, SUPPORTED_CHAPTERS
from image_analyzer import analyze_image_ensemble, predict_chapters, analyze_and_predict
from category_lookup import get_extra_keywords
import analysis_cache
from cpsc_hts import check_cpsc

DATA_DIR = Path(__file__).parent / "data"
_USITC_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "*/*",
    "Referer": "https://hts.usitc.gov/",
}

AUTO_KEY = "__auto__"
MAX_BATCH = 10


def _download_chapter(chapter_key: str, data_file: str) -> bool:
    chapter_str = str(int(chapter_key)).zfill(2)
    try:
        ranges_url = f"https://hts.usitc.gov/reststop/ranges?docNumber={chapter_str}"
        req = urllib.request.Request(ranges_url, headers=_USITC_HEADERS)
        with urllib.request.urlopen(req, timeout=15) as resp:
            rng = json.loads(resp.read().decode("utf-8"))

        export_url = (
            f"https://hts.usitc.gov/reststop/exportList"
            f"?from={rng['Starting_Number']}&to={rng['Ending_Number']}&format=JSON&styles=true"
        )
        req = urllib.request.Request(export_url, headers=_USITC_HEADERS)
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        out_path = DATA_DIR / data_file
        out_path.parent.mkdir(exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        st.error(f"ダウンロードに失敗しました: {e}")
        return False


LOGS_DIR = Path(__file__).parent / "logs"
LOGS_DIR.mkdir(exist_ok=True)
FEEDBACK_LOG = LOGS_DIR / "feedback.csv"

JP_LABELS_PATH = Path(__file__).parent / "jp_labels.json"
with open(JP_LABELS_PATH, encoding="utf-8") as f:
    JP_LABELS = json.load(f)

# ── ページ設定 ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="HTSコード判定支援ツール", layout="wide")
st.markdown("""
<style>
/* 判定結果カードのExpanderタイトル（📦で始まるもの）を大きく・色付きに */
div[data-testid="stExpander"] > details > summary p:first-child {
    font-size: 1.35rem !important;
    font-weight: 700 !important;
    color: #1a6eb5 !important;
}
/* 内側の小Expander（画像解析結果・上位分類）は通常サイズに戻す */
div[data-testid="stExpander"] div[data-testid="stExpander"] > details > summary p {
    font-size: 0.9rem !important;
    font-weight: 400 !important;
    color: inherit !important;
}
</style>
""", unsafe_allow_html=True)
st.title("米国HTSコード判定支援ツール")
st.caption("商品画像を主判断とし、テキスト情報を補助としてHTSコード候補を提示します（最大10件同時判定）。最終判断は担当者が行ってください。")

# モックモード警告（API課金ゼロのテスト用）
from config import MOCK_MODE as _MOCK_MODE, MOCK_EXPLICIT as _MOCK_EXPLICIT
if _MOCK_MODE:
    _reason = "Secrets/環境変数で有効化" if _MOCK_EXPLICIT else "APIキー未設定のため自動でモックに切替"
    st.warning(
        f"🧪 **モックモード稼働中**（{_reason}） — Claude APIを呼び出していません（課金ゼロ）。"
        "判定結果は補足テキストから生成したダミーで、精度は参考になりません。"
        "UI・バッチ処理・Excel出力など動作確認用です。"
    )

# サイドバー: キャッシュ統計
with st.sidebar:
    st.subheader("📊 キャッシュ統計（過去30日）")
    stats = analysis_cache.get_stats()
    st.metric("総リクエスト数", f"{stats['total_requests']:,}")
    col_l1, col_l2 = st.columns(2)
    col_l1.metric("L1ヒット率", f"{stats['hit_rate_l1']:.1%}", help="同一画像キャッシュ（APIスキップ）")
    col_l2.metric("L2ヒット率", f"{stats['hit_rate_l2']:.1%}", help="同じ分析結果キャッシュ（照合スキップ）")
    st.metric("推定削減コスト（API）", f"${stats['saved_usd_30d']:,.2f}")
    st.caption(f"キャッシュ済み画像: {stats['cached_images']:,} 件　分析パターン: {stats['cached_analyses']:,} 種")

    with st.expander("🧹 キャッシュ管理"):
        st.caption(
            "分類ルールを更新した後、古い判定結果が残る場合はL2をクリアしてください。"
            "L2クリアは再判定でAPI課金は発生しません（画像解析のL1は保持）。"
        )
        if st.button("L2クリア（HTS結果・再判定／API不要）", use_container_width=True):
            n = analysis_cache.clear_hts_cache()
            analysis_cache.clear_chapter_cache()
            st.success(f"L2と章キャッシュをクリアしました（HTS結果 {n} 件）。次回は最新ルールで再判定します。")
        st.caption("⚠️ 下は画像解析(L1)も消去 → 再判定でAPI再課金が発生します。")
        if st.button("全キャッシュをクリア（L1含む・要API再課金）", use_container_width=True):
            r = analysis_cache.clear_all()
            st.warning(f"全キャッシュを削除しました（L1 {r['l1']}・L2 {r['l2']}・章 {r['chapters']} 件）。")

# ── セッション初期化 ────────────────────────────────────────────────────────
if "batch_results" not in st.session_state:
    st.session_state["batch_results"] = []
if "form_version" not in st.session_state:
    st.session_state["form_version"] = 0

v = st.session_state["form_version"]

top_cols = st.columns([4, 1])
with top_cols[1]:
    if st.button("🔄 クリア"):
        st.session_state["form_version"] += 1
        st.session_state["batch_results"] = []
        st.rerun()

# ── 章選択 ──────────────────────────────────────────────────────────────────
chapter_options = [AUTO_KEY] + list(SUPPORTED_CHAPTERS.keys())

def _chapter_label(k):
    if k == AUTO_KEY:
        return "🔍 自動判定（章を自動で推定）"
    return SUPPORTED_CHAPTERS[k]["label"]

chapter_key = st.selectbox(
    "対象Chapter",
    options=chapter_options,
    format_func=_chapter_label,
    key=f"chapter_key_{v}",
)

if chapter_key != AUTO_KEY:
    chapter_conf = SUPPORTED_CHAPTERS[chapter_key]
    chapter_data_path = DATA_DIR / chapter_conf["data_file"]
    if not chapter_data_path.exists():
        st.warning(
            f"**{chapter_conf['label']}** のHTSデータがまだダウンロードされていません。"
            "下のボタンでUSITCから取得できます（初回のみ）。"
        )
        if st.button("📥 HTSデータをダウンロード", type="primary"):
            with st.spinner("USITCからデータを取得しています..."):
                ok = _download_chapter(chapter_key, chapter_conf["data_file"])
            if ok:
                st.success("ダウンロード完了。判定を開始できます。")
                time.sleep(1)
                st.rerun()
        st.stop()

# ── 入力フォーム ─────────────────────────────────────────────────────────────
st.subheader("商品情報（テキスト補助・全件共通）")
tcols = st.columns(4)
product_name  = tcols[0].text_input("品名", key=f"product_name_{v}")
material      = tcols[1].text_input("材質", key=f"material_{v}")
function_desc = tcols[2].text_input("用途・機能", key=f"function_desc_{v}")
spec          = tcols[3].text_input("仕様・補足", key=f"spec_{v}")
text_context  = " ".join(p for p in [product_name, material, function_desc, spec] if p)

st.subheader(f"商品画像（最大{MAX_BATCH}件）")
uploaded_images = st.file_uploader(
    "画像をアップロード（複数選択可・最大10件）",
    type=["png", "jpg", "jpeg"],
    accept_multiple_files=True,
    key=f"uploaded_images_{v}",
)

if uploaded_images:
    if len(uploaded_images) > MAX_BATCH:
        st.warning(f"最大{MAX_BATCH}件まで処理できます。先頭{MAX_BATCH}件を使用します。")
        uploaded_images = uploaded_images[:MAX_BATCH]

    # プレビュー（最大5列）
    cols_per_row = min(len(uploaded_images), 5)
    preview_cols = st.columns(cols_per_row)
    for i, img in enumerate(uploaded_images):
        preview_cols[i % cols_per_row].image(img, caption=img.name, width='stretch')

# ── ヘルパー関数 ─────────────────────────────────────────────────────────────
_HINT_EXCLUDE_CHAPTERS: dict[str, list[str]] = {
    "cutting mat": ["84", "95"],
    "ruler": ["84", "95"],
    "staples": ["95"],
    "stapler": ["95"],
}

# 画像ヒントに応じて候補章を強制的に追加する（誤った章のみに誘導されるのを防ぐ）。
# 例: バッジはプラ製→39 / 金属製→83 のどちらにもなり得るため両方を候補に含める。
_HINT_INCLUDE_CHAPTERS: dict[str, list[str]] = {
    "badge": ["39", "83", "71"],
    "pin badge": ["39", "83", "71"],
    "brooch": ["71", "39", "83"],
    "trading card": ["95", "49"],
    "game card": ["95", "49"],
    "playing card": ["95", "49"],
    "game console": ["95", "85"],
    "gaming device": ["95", "85"],
    "handheld game": ["95", "85"],
    "arcade": ["95", "85"],
    "recorded dvd": ["85"],
    "movie dvd": ["85"],
    "music cd": ["85"],
    "wrench": ["82"],
    "spanner": ["82"],
    "hand tool": ["82"],
    "pliers": ["82"],
    "screwdriver": ["82"],
    "sneaker": ["64"],
    "athletic shoe": ["64"],
    "running shoe": ["64"],
    "sports shoe": ["64"],
    "shoe": ["64"],
    "footwear": ["64"],
    "rubber stamp": ["96"],
    "stamp": ["96"],
    "pressed flower": ["67"],
    "dried flower": ["67"],
    "artificial flower": ["67"],
    "woven cotton fabric": ["52"],
    "printed cotton fabric": ["52"],
    "cotton fabric": ["52"],
    "fabric": ["52", "61", "62", "63"],
    "t-shirt": ["61"],
    "trousers": ["62"],
    "denim": ["62"],
    "jeans": ["62"],
    "shorts": ["62"],
    "charm": ["39", "71"],
    "steering wheel cover": ["87"],
    # アニメ・ホビー
    "figure toy": ["95", "39"],
    "action figure": ["95"],
    "toy figure": ["95"],
    "collectible figure": ["95"],
    "building block": ["95"],
    "construction toy": ["95"],
    "block toy": ["95"],
    "nano block": ["95"],
    "micro block": ["95"],
    "assembly toy": ["95"],
    "stuffed toy": ["95"],
    "plush": ["95"],
    "dakimakura": ["63"],
    "tapestry": ["63"],
    "keychain": ["39", "83"],
    "gashapon": ["95"],
    "capsule toy": ["95"],
    "clear folder": ["39", "48"],
    "mug": ["69"],
    "towel": ["63"],
    "cushion": ["94"],
    "mouse pad": ["39"],
    # 帽子・アクセサリー
    "hat": ["65"],
    "cap": ["65"],
    "headwear": ["65"],
    "wig": ["67"],
    "necklace": ["71"],
    "ring": ["71"],
    "earring": ["71"],
    "bracelet": ["71"],
    # 家電
    "smartphone": ["85"],
    "camera": ["90"],
    "laptop": ["84"],
    "earphone": ["85"],
    "headphone": ["85"],
    "speaker": ["85"],
    # 食器・インテリア
    "ceramic": ["69"],
    "tableware": ["69"],
    "lighting": ["94"],
    # 伝統ゲーム
    "board game": ["95"],
    "card game": ["95"],
    "magazine": ["49"],
    "periodical": ["49"],
    "printed magazine": ["49"],
    "hobby magazine": ["49"],
    "journal": ["49"],
    "publication": ["49"],
    "radio controlled": ["85", "95"],
    "rc controller": ["85"],
    "radio remote control": ["85"],
    "die-cast car": ["95"],
    "die-cast": ["95"],
    "scale model car": ["95"],
    "toy vehicle": ["95"],
    "collectible car": ["95"],
    "children's book": ["49"],
    "illustrated book": ["49"],
    "textbook": ["49"],
    "schoolbook": ["49"],
    "kitchen knife": ["82"],
    "chef knife": ["82"],
    "butcher knife": ["82"],
    "tool case": ["42"],
    "equipment case": ["42"],
    "calligraphy": ["49"],
    "wall art": ["49"],
    "art print": ["49"],
    "pet bed": ["63"],
    "pet mat": ["63"],
    "pet cushion": ["63"],
    "storage box": ["39"],
    "storage organizer": ["39"],
    "desk organizer": ["39"],
    "drawer organizer": ["39"],
    "organizer box": ["39"],
    # バッグ
    "handbag": ["42"],
    "backpack": ["42"],
    "wallet": ["42"],
    "wristwatch": ["91"],
    "vinyl record": ["85"],
}


# 画像解析の表示用 英語→日本語 用語辞書（モデルのJA出力が無い場合のフォールバック）
_TERM_JA: dict[str, str] = {
    "plastic": "プラスチック", "metal": "金属", "wood": "木", "wooden": "木製",
    "paper": "紙", "cardboard": "段ボール", "paperboard": "板紙", "rubber": "ゴム",
    "glass": "ガラス", "ceramic": "陶磁器", "porcelain": "磁器", "textile": "繊維",
    "cotton": "綿", "polyester": "ポリエステル", "steel": "鋼", "stainless steel": "ステンレス鋼",
    "iron": "鉄", "aluminum": "アルミニウム", "aluminium": "アルミニウム", "brass": "真鍮",
    "leather": "革", "silicone": "シリコーン", "vinyl": "ビニール", "acrylic": "アクリル",
    "fabric": "布", "nylon": "ナイロン", "pvc": "塩化ビニル", "foam": "発泡材",
    "box": "箱", "packaging box": "包装箱", "cardboard box": "段ボール箱",
    "folding carton": "折りたたみ式カートン", "folding carton box": "折りたたみ式カートン箱",
    "retail packaging": "小売用包装", "packaging": "包装",
    "product packaging and storage": "製品の包装・保管",
    "pin badge": "ピンバッジ", "badge": "バッジ", "brooch": "ブローチ",
    "toy": "玩具", "doll": "人形", "figure": "フィギュア", "tricycle": "三輪車",
    "book": "書籍", "printed book": "印刷書籍", "card": "カード",
}


def _ja_term(text: str) -> str:
    """用語辞書で英語→日本語に変換。未収録はそのまま返す。"""
    if not text:
        return text
    return _TERM_JA.get(text.lower().strip(), text)


def _analysis_ja_view(analysis: dict) -> dict:
    """画像解析結果を日本語表示用のdictに整える。
    モデルが返したJAフィールドを優先し、無ければ用語辞書でフォールバック。
    """
    def pick(en_key, ja_key):
        return analysis.get(ja_key) or _ja_term(analysis.get(en_key, ""))

    kws_ja = analysis.get("keywords_ja")
    if not kws_ja:
        kws_ja = [_ja_term(k) for k in analysis.get("keywords", [])]
    return {
        "材質": pick("material", "material_ja"),
        "用途・機能": pick("function", "function_ja"),
        "商品区分": pick("category_hint", "category_hint_ja"),
        "キーワード": kws_ja,
    }


import re as _re
# 採用コードの表示で、それ自体では商品が分からない汎用語（その他・素材名のみ）。
# これらの場合は full_description を上位へ遡って商品の種類を表示する。
_GENERIC_DESC = {
    "other", "other, including parts", "",
    "of cotton", "of leather", "of plastics", "of rubber", "of wool",
    "of man-made fibers", "of synthetic fibers", "of artificial fibers",
    "of other textile materials", "of paper", "of metal", "of wood",
    "of vegetable fibers", "of textile materials", "parts",
}


def _is_generic_desc(s: str) -> bool:
    # 末尾の国内細分コード "(347)" 等と記号を除去して判定
    t = _re.sub(r"\(\d+\)\s*$", "", s).rstrip(":").strip().lower()
    return t in _GENERIC_DESC


def _entry_label_ja(entry: dict | None) -> str:
    """採用コードの表示ラベル（日本語優先・汎用語/素材名のみを回避）。
    ① override由来の専用表示名 → ② 説明が汎用語なら上位見出し → ③ JP_LABELSで日本語化。
    """
    if not entry:
        return ""
    if entry.get("_label_ja"):
        return entry["_label_ja"]
    desc = (entry.get("description") or "").rstrip(":").strip()
    if _is_generic_desc(desc):
        segs = [s.rstrip(":").strip() for s in (entry.get("full_description") or "").split(" > ")]
        for s in reversed(segs):
            if s and not _is_generic_desc(s):
                desc = s
                break
    return JP_LABELS.get(_re.sub(r"\(\d+\)\s*$", "", desc).rstrip(":").strip(), "") or desc


# 日本語カテゴリ語→英語のマッピング（classifier._JA_TO_EN_HINT と同内容）
# app.py内でモジュールレベルに定義することで、インポート順序問題を回避する
_JA_TO_EN_APP: dict[str, str] = {
    "ショーツ": "shorts", "短パン": "shorts", "ハーフパンツ": "shorts",
    "ズボン": "trousers", "パンツ": "trousers", "スラックス": "trousers",
    "ジーンズ": "jeans", "デニム": "denim",
    "tシャツ": "t-shirt", "ｔシャツ": "t-shirt",
    "スニーカー": "sneaker", "運動靴": "athletic shoe", "スポーツシューズ": "sports shoe",
    "ピンバッジ": "pin badge", "缶バッジ": "button badge", "バッジ": "badge",
    "アクリルスタンド": "acrylic stand", "アクスタ": "acrylic stand",
    "チャーム": "charm", "キーチェーン": "charm",
    "トレーディングカード": "trading card", "トレカ": "trading card",
    "ゲーム機": "game console", "ゲームソフト": "game console",
    "テレビ": "television", "テレビ受像機": "television",
    "dvd": "recorded dvd", "ブルーレイ": "recorded blu-ray",
    "同人誌": "doujinshi", "同人": "self-published",
    "漫画": "manga", "コミック": "comic book",
    "ゴム印": "rubber stamp", "スタンプ": "rubber stamp",
    "造花": "artificial flower", "プレスフラワー": "pressed flower",
    "ドライフラワー": "dried flower",
    "レンチ": "wrench", "スパナ": "spanner",
    "綿生地": "woven cotton fabric", "プリント生地": "printed cotton fabric",
    "綿布": "woven cotton fabric",
    "ステアリングホイールカバー": "steering wheel cover",
    "ステアリングカバー": "steering wheel cover",
    "カッティングマット": "cutting mat",
    "収納ボックス": "storage box",
    "整理ボックス": "drawer organizer",
    "オーガナイザー": "desk organizer",
    "ラジコン": "radio controlled rc toy remote control",
    "プロポ": "radio control transmitter rc controller",
    "ラジコン送信機": "radio remote control transmitter",
    "ダイキャストカー": "die-cast car scale model toy vehicle",
    "ダイキャスト": "die-cast scale model",
    "玩具車両": "toy vehicle scale model car",
    "レーシングカーモデル": "scale model racing car toy",
    "コレクタブルカー": "collectible car scale model",
    "児童書": "children's book illustrated book",
    "挿絵本": "illustrated children's book",
    "教科書": "textbook schoolbook",
    "歴史教材": "history textbook educational book",
    "包丁": "kitchen knife chef knife",
    "キッチンナイフ": "kitchen knife",
    "シェフナイフ": "chef knife kitchen knife",
    "調理用ナイフ": "kitchen knife butcher knife",
    "工具ケース": "tool case carrying case",
    "器具ケース": "tool case equipment case",
    "書道": "calligraphy art print",
    "書道アート": "calligraphy art print decorative",
    "壁掛けアート": "wall art print decorative",
    "アートプリント": "art print decorative picture",
    "ペットベッド": "pet bed textile mat",
    "ペット用ベッド": "pet bed textile mat",
    "ペットマット": "pet mat cushion",
    "クッションマット": "cushion mat textile",
    "雑誌": "magazine periodical",
    "印刷雑誌": "printed magazine periodical",
    "ホビー雑誌": "hobby magazine periodical",
    "定期刊行物": "periodical journal",
    "出版物": "publication periodical",
    "ナノブロック": "nano block building block",
    "マイクロブロック": "micro block building block",
    "ブロック玩具": "building block toy",
    "組み立て玩具": "construction toy assembly toy",
    "ブロックフィギュア": "block toy figure",
    "アクションフィギュア": "action figure",
    "コレクタブルフィギュア": "collectible figure",
    "フィギュア": "toy figure", "ぬいぐるみ": "stuffed toy",
    "抱き枕": "dakimakura", "タペストリー": "tapestry",
    "キーホルダー": "keychain", "ガチャ": "gashapon",
    "帽子": "hat", "ウィッグ": "wig",
    "ネックレス": "necklace", "リング": "ring",
    "イヤリング": "earring", "ピアス": "earring",
    "腕時計": "wristwatch", "食器": "tableware",
    "マグカップ": "mug", "タオル": "towel",
    "ハンドバッグ": "handbag", "リュック": "backpack", "財布": "wallet",
    "スマートフォン": "smartphone", "カメラ": "camera",
    "レコード": "vinyl record",
    "花札": "card game", "将棋": "board game", "囲碁": "board game",
    "風呂敷": "furoshiki wrapping cloth",
    "ラッピングクロス": "wrapping cloth",
    "パッチワーク": "patchwork fabric",
    "装飾用織物": "decorative cloth",
    "装飾的な布地": "decorative cloth",
}


def _normalize_hint_app(text: str) -> str:
    """日本語カテゴリ語を英語に正規化して返す（app.py専用、importなし）。"""
    result = text.lower()
    for ja, en in _JA_TO_EN_APP.items():
        if ja.lower() in result:
            result += f" {en}"
    return result


# 品名にこれらの代表語が含まれていたら、画像解析API(高コスト)を呼ばず
# テキストだけで分類する。値: category_hint(英)・候補章・既定材質(任意)。
# 確実に分類が定まる品目（オーバーライドがある品目）に限定する。
_DECISIVE_TERMS: dict[str, dict] = {
    "ピンバッジ": {"hint": "pin badge", "chapters": ["39", "83", "71"], "material": "plastic"},
    "缶バッジ": {"hint": "button badge", "chapters": ["39", "83", "71"], "material": "plastic"},
    "カッティングマット": {"hint": "cutting mat", "chapters": ["39"], "material": "plastic"},
    "アクリルスタンド": {"hint": "acrylic stand", "chapters": ["39"], "material": "plastic"},
    "アクスタ": {"hint": "acrylic stand", "chapters": ["39"], "material": "plastic"},
    "チャーム": {"hint": "charm", "chapters": ["39", "71"], "material": "plastic"},
    "トレカ": {"hint": "trading card", "chapters": ["95", "49"]},
    "トレーディングカード": {"hint": "trading card", "chapters": ["95", "49"]},
    "ゲーム機": {"hint": "game console", "chapters": ["95", "85"]},
    "テレビ": {"hint": "television receiver", "chapters": ["85"]},
    "ブルーレイ": {"hint": "recorded blu-ray", "chapters": ["85"]},
    "音楽cd": {"hint": "music cd", "chapters": ["85"]},
    "同人誌": {"hint": "self-published book", "chapters": ["49"], "material": "paper"},
    "漫画": {"hint": "comic book", "chapters": ["49"], "material": "paper"},
    "コミック": {"hint": "comic book", "chapters": ["49"], "material": "paper"},
    "tシャツ": {"hint": "t-shirt", "chapters": ["61"], "material": "cotton"},
    "ｔシャツ": {"hint": "t-shirt", "chapters": ["61"], "material": "cotton"},
    "ズボン": {"hint": "trousers", "chapters": ["62"], "material": "cotton"},
    "ジーンズ": {"hint": "jeans", "chapters": ["62"], "material": "cotton"},
    "デニム": {"hint": "denim", "chapters": ["62"], "material": "cotton"},
    "ショーツ": {"hint": "shorts", "chapters": ["62"], "material": "cotton"},
    "スニーカー": {"hint": "sneaker", "chapters": ["64"]},
    "レンチ": {"hint": "wrench", "chapters": ["82"], "material": "metal"},
    "スパナ": {"hint": "spanner", "chapters": ["82"], "material": "metal"},
    "ゴム印": {"hint": "rubber stamp", "chapters": ["96"]},
    "造花": {"hint": "artificial flower", "chapters": ["67"]},
    "プレスフラワー": {"hint": "pressed flower", "chapters": ["67", "39"]},
    "ドライフラワー": {"hint": "dried flower", "chapters": ["67", "39"]},
    "綿生地": {"hint": "woven cotton fabric", "chapters": ["52"], "material": "cotton"},
    "プリント生地": {"hint": "printed cotton fabric", "chapters": ["52"], "material": "cotton"},
    "綿布": {"hint": "woven cotton fabric", "chapters": ["52"], "material": "cotton"},
    "生地": {"hint": "cotton fabric", "chapters": ["52", "61", "62", "63"]},
    "風呂敷": {"hint": "furoshiki wrapping cloth", "chapters": ["63"], "material": "cotton"},
    "ラッピングクロス": {"hint": "wrapping cloth decorative", "chapters": ["63", "52"]},
    "パッチワーク": {"hint": "patchwork fabric", "chapters": ["52", "63"]},

    # ── アニメグッズ・ホビー（駿河屋/Neokyo メインカテゴリ） ─────────────────
    "フィギュア": {"hint": "figure toy", "chapters": ["95", "39"]},
    "ぬいぐるみ": {"hint": "stuffed toy plush", "chapters": ["95"]},
    "抱き枕": {"hint": "dakimakura pillow cover", "chapters": ["63"]},
    "タペストリー": {"hint": "tapestry textile wall hanging", "chapters": ["63"]},
    "色紙": {"hint": "autograph board illustrated card", "chapters": ["49"]},
    "クリアファイル": {"hint": "clear plastic folder", "chapters": ["39", "48"]},
    "缶ケース": {"hint": "metal tin case", "chapters": ["83"]},
    "キーホルダー": {"hint": "keychain keyring", "chapters": ["39", "83", "71"]},
    "ストラップ": {"hint": "strap accessory", "chapters": ["39", "83"]},
    "スタンドポップ": {"hint": "acrylic stand pop display", "chapters": ["39"]},
    "ポスター": {"hint": "printed poster", "chapters": ["49"]},
    "ブロマイド": {"hint": "printed photo card bromide", "chapters": ["49"]},
    "フォトカード": {"hint": "printed photo card", "chapters": ["49"]},
    "缶バッチ": {"hint": "button badge", "chapters": ["39", "83"]},
    "ラバーストラップ": {"hint": "rubber strap charm", "chapters": ["39"]},
    "マスコット": {"hint": "mascot figure toy", "chapters": ["95", "39"]},
    "コースター": {"hint": "coaster", "chapters": ["39", "48", "63"]},
    "マグカップ": {"hint": "mug ceramic cup", "chapters": ["69"]},
    "タオル": {"hint": "towel textile", "chapters": ["63"]},
    "手ぬぐい": {"hint": "tenugui cotton hand towel", "chapters": ["63"]},
    "スマホケース": {"hint": "smartphone phone case cover", "chapters": ["39", "85"]},
    "マウスパッド": {"hint": "mouse pad", "chapters": ["39", "85"]},
    "クッション": {"hint": "cushion pillow", "chapters": ["94"]},
    "ランダムグッズ": {"hint": "random character goods toy", "chapters": ["95", "39"]},
    "ガチャ": {"hint": "capsule toy gashapon", "chapters": ["95"]},
    "ガチャガチャ": {"hint": "capsule toy gashapon", "chapters": ["95"]},
    "カプセルトイ": {"hint": "capsule toy", "chapters": ["95"]},
    "ミニフィギュア": {"hint": "mini figure toy", "chapters": ["95"]},
    "プライズフィギュア": {"hint": "prize figure toy", "chapters": ["95"]},
    "スケールフィギュア": {"hint": "scale figure collectible", "chapters": ["95", "39"]},

    # ── ゲーム（駿河屋/Neokyo Gaming） ─────────────────────────────────────
    "ゲームソフト": {"hint": "video game software cartridge", "chapters": ["85", "95"]},
    "Switch": {"hint": "nintendo switch game console", "chapters": ["95", "85"]},
    "ニンテンドースイッチ": {"hint": "nintendo switch game console", "chapters": ["95", "85"]},
    "プレイステーション": {"hint": "playstation game console", "chapters": ["95", "85"]},
    "PlayStation": {"hint": "playstation game console", "chapters": ["95", "85"]},
    "ニンテンドーDS": {"hint": "nintendo ds handheld game console", "chapters": ["95", "85"]},
    "ゲームボーイ": {"hint": "game boy handheld game console", "chapters": ["95", "85"]},
    "ファミコン": {"hint": "famicom retro video game console", "chapters": ["95", "85"]},
    "アーケード": {"hint": "arcade game machine", "chapters": ["95", "85"]},
    "コントローラー": {"hint": "game controller joystick", "chapters": ["95", "85"]},

    # ── 音楽・映像メディア ─────────────────────────────────────────────────
    "cd": {"hint": "music cd", "chapters": ["85"]},
    "dvd": {"hint": "recorded dvd", "chapters": ["85"]},
    "blu-ray": {"hint": "recorded blu-ray", "chapters": ["85"]},
    "レコード": {"hint": "vinyl record", "chapters": ["85"]},
    "カセット": {"hint": "cassette tape", "chapters": ["85"]},
    "vhs": {"hint": "vhs tape video cassette", "chapters": ["85"]},
    "写真集": {"hint": "photo book printed publication", "chapters": ["49"]},
    "楽譜": {"hint": "sheet music printed", "chapters": ["49"]},

    # ── 本・書籍（Neokyo Book） ─────────────────────────────────────────────
    "小説": {"hint": "novel printed book", "chapters": ["49"], "material": "paper"},
    "雑誌": {"hint": "magazine periodical", "chapters": ["49"], "material": "paper"},
    "画集": {"hint": "art book illustrated book", "chapters": ["49"], "material": "paper"},
    "攻略本": {"hint": "game strategy guide book", "chapters": ["49"], "material": "paper"},
    "絵本": {"hint": "picture book", "chapters": ["49"], "material": "paper"},

    # ── ファッション（Neokyo Fashion / メルカリ） ────────────────────────────
    "ワンピース": {"hint": "dress women's garment", "chapters": ["61", "62"]},
    "スカート": {"hint": "skirt women's garment", "chapters": ["61", "62"]},
    "ブラウス": {"hint": "blouse shirt women's", "chapters": ["61", "62"]},
    "ニット": {"hint": "knitted sweater", "chapters": ["61"]},
    "セーター": {"hint": "sweater knit", "chapters": ["61"]},
    "パーカー": {"hint": "hoodie sweatshirt", "chapters": ["61"]},
    "コート": {"hint": "coat outerwear", "chapters": ["61", "62"]},
    "ジャケット": {"hint": "jacket outerwear", "chapters": ["61", "62"]},
    "スーツ": {"hint": "suit jacket trousers", "chapters": ["61", "62"]},
    "帽子": {"hint": "hat cap headwear", "chapters": ["65"]},
    "マフラー": {"hint": "scarf muffler", "chapters": ["61", "62"]},
    "手袋": {"hint": "gloves", "chapters": ["61", "62"]},
    "ソックス": {"hint": "socks hosiery", "chapters": ["61"]},
    "靴下": {"hint": "socks hosiery", "chapters": ["61"]},
    "パンプス": {"hint": "pumps women's footwear heels", "chapters": ["64"]},
    "サンダル": {"hint": "sandals footwear", "chapters": ["64"]},
    "ブーツ": {"hint": "boots footwear", "chapters": ["64"]},
    "革靴": {"hint": "leather shoes footwear", "chapters": ["64"]},
    "ウィッグ": {"hint": "wig hair accessory", "chapters": ["67"]},
    "コスプレ": {"hint": "cosplay costume", "chapters": ["61", "62", "95"]},
    "着物": {"hint": "kimono traditional japanese garment", "chapters": ["62"]},
    "浴衣": {"hint": "yukata japanese garment", "chapters": ["62"]},
    "ベルト": {"hint": "belt leather", "chapters": ["42", "62"]},

    # ── バッグ・小物 ────────────────────────────────────────────────────────
    "ハンドバッグ": {"hint": "handbag", "chapters": ["42"]},
    "リュック": {"hint": "backpack rucksack", "chapters": ["42"]},
    "ショルダーバッグ": {"hint": "shoulder bag", "chapters": ["42"]},
    "トートバッグ": {"hint": "tote bag", "chapters": ["42"]},
    "財布": {"hint": "wallet purse leather", "chapters": ["42"]},
    "ポーチ": {"hint": "pouch small bag", "chapters": ["42", "39"]},
    "サコッシュ": {"hint": "sacoche sling bag", "chapters": ["42"]},

    # ── 時計・ジュエリー ────────────────────────────────────────────────────
    "腕時計": {"hint": "wristwatch", "chapters": ["91"]},
    "置き時計": {"hint": "table clock", "chapters": ["91"]},
    "ネックレス": {"hint": "necklace jewelry", "chapters": ["71"]},
    "収納ボックス": {"hint": "storage box organizer", "chapters": ["39"]},
    "整理ボックス": {"hint": "drawer organizer storage box", "chapters": ["39"]},
    "オーガナイザー": {"hint": "desk organizer storage", "chapters": ["39"]},
    "ラジコン送信機": {"hint": "radio remote control transmitter hobby", "chapters": ["85"]},
    "プロポ": {"hint": "radio control transmitter rc hobby", "chapters": ["85"]},
    "ダイキャストカー": {"hint": "die-cast car scale model toy vehicle", "chapters": ["95"]},
    "包丁": {"hint": "kitchen knife chef knife", "chapters": ["82"]},
    "キッチンナイフ": {"hint": "kitchen knife", "chapters": ["82"]},
    "書道": {"hint": "calligraphy art print decorative", "chapters": ["49"]},
    "ペットベッド": {"hint": "pet bed textile mat", "chapters": ["63"]},
    "ペット用ベッド": {"hint": "pet bed textile mat", "chapters": ["63"]},
    "ホビー雑誌": {"hint": "hobby magazine periodical", "chapters": ["49"]},
    "ナノブロック": {"hint": "nano block building block toy", "chapters": ["95"]},
    "マイクロブロック": {"hint": "micro block building block toy", "chapters": ["95"]},
    "ブロック玩具": {"hint": "building block construction toy", "chapters": ["95"]},
    "組み立て玩具": {"hint": "construction toy assembly plastic toy", "chapters": ["95"]},
    "アクションフィギュア": {"hint": "action figure toy", "chapters": ["95"]},
    "コレクタブルフィギュア": {"hint": "collectible figure toy", "chapters": ["95"]},
    "ステアリングホイールカバー": {"hint": "steering wheel cover", "chapters": ["87"]},
    "リング": {"hint": "ring jewelry", "chapters": ["71"]},
    "イヤリング": {"hint": "earring jewelry", "chapters": ["71"]},
    "ピアス": {"hint": "piercing earring jewelry", "chapters": ["71"]},
    "ブレスレット": {"hint": "bracelet jewelry", "chapters": ["71"]},
    "指輪": {"hint": "ring jewelry", "chapters": ["71"]},

    # ── 家電・電子機器 ────────────────────────────────────────────────────
    "スマートフォン": {"hint": "smartphone mobile phone", "chapters": ["85"]},
    "タブレット": {"hint": "tablet computer", "chapters": ["84", "85"]},
    "ノートパソコン": {"hint": "laptop notebook computer", "chapters": ["84"]},
    "イヤホン": {"hint": "earphone headphone audio", "chapters": ["85"]},
    "ヘッドフォン": {"hint": "headphone audio", "chapters": ["85"]},
    "スピーカー": {"hint": "speaker audio", "chapters": ["85"]},
    "カメラ": {"hint": "camera optical", "chapters": ["90"]},
    "プリンター": {"hint": "printer machine", "chapters": ["84"]},
    "電子レンジ": {"hint": "microwave oven", "chapters": ["85"]},
    "炊飯器": {"hint": "rice cooker electric", "chapters": ["85"]},
    "掃除機": {"hint": "vacuum cleaner", "chapters": ["85"]},
    "ドライヤー": {"hint": "hair dryer electric", "chapters": ["85"]},

    # ── 伝統ゲーム（Neokyo Traditional Games） ───────────────────────────
    "花札": {"hint": "hanafuda card game", "chapters": ["95"]},
    "将棋": {"hint": "shogi board game", "chapters": ["95"]},
    "囲碁": {"hint": "go board game", "chapters": ["95"]},
    "けん玉": {"hint": "kendama skill toy", "chapters": ["95"]},
    "かるた": {"hint": "karuta card game", "chapters": ["95"]},
    "ボードゲーム": {"hint": "board game", "chapters": ["95"]},
    "カードゲーム": {"hint": "card game", "chapters": ["95"]},
    "マジック": {"hint": "magic trick toy", "chapters": ["95"]},

    # ── インテリア・生活雑貨 ─────────────────────────────────────────────
    "食器": {"hint": "tableware ceramic", "chapters": ["69"]},
    "湯呑": {"hint": "teacup ceramic", "chapters": ["69"]},
    "お椀": {"hint": "bowl ceramic", "chapters": ["69"]},
    "照明": {"hint": "lighting lamp", "chapters": ["94"]},
    "スタンドライト": {"hint": "floor lamp lighting", "chapters": ["94"]},
    "額縁": {"hint": "picture frame", "chapters": ["83", "39", "44"]},
}


def _decisive_match(text: str) -> tuple[str, dict] | tuple[None, None]:
    """品名テキストに代表語が含まれていれば (代表語, 情報) を返す。
    長い語を先に照合し、短い語が長い語の一部に誤マッチするのを防ぐ。
    例: 「リング」が「ステアリング」に誤マッチしないよう、
        「ステアリングホイールカバー」を「リング」より先に評価する。
    """
    t = (text or "").lower()
    # 長い語を優先（長さ降順でソート）
    for term in sorted(_DECISIVE_TERMS, key=len, reverse=True):
        if term.lower() in t:
            return term, _DECISIVE_TERMS[term]
    return None, None


def _classify_text_only(image_bytes: bytes, image_name: str, text_ctx: str,
                        term: str, info: dict, suruga_kw: list[str]) -> dict:
    """代表語ヒット時：画像APIを呼ばずテキストだけで分類する。"""
    mat = info.get("material", "")  # 既定材質（品名からは素材が取れないため）
    query = {
        "product_name": "", "material": mat, "category_hint": info["hint"],
        "function": "", "keywords": [info["hint"]] + suruga_kw, "spec": "",
    }
    chapters = [c for c in info["chapters"] if c in SUPPORTED_CHAPTERS]
    for k in chapters:
        if not (DATA_DIR / SUPPORTED_CHAPTERS[k]["data_file"]).exists():
            _download_chapter(k, SUPPORTED_CHAPTERS[k]["data_file"])
    chapter_files = [
        (k, SUPPORTED_CHAPTERS[k]["data_file"]) for k in chapters
        if (DATA_DIR / SUPPORTED_CHAPTERS[k]["data_file"]).exists()
    ]
    results = apply_hts_overrides(
        classify_per_chapter_ensemble([query], chapter_files=chapter_files, top_n_per_chapter=3),
        [query],
    )
    return {
        "filename": image_name, "image_bytes": image_bytes, "image_analysis": None,
        "results": results, "detected_chapters": chapters, "is_auto": True,
        "cache_hit_l1": False, "cache_hit_l2": False,
        "text_only": True, "decisive_term": term,
    }


def _build_query(analysis: dict | None, suruga_keywords: list[str]) -> dict:
    if analysis:
        return {
            "product_name": "",
            "material": analysis.get("material", ""),
            "category_hint": analysis.get("category_hint", ""),
            "keywords": analysis.get("keywords", []) + suruga_keywords,
            "function": analysis.get("function", ""),
            "spec": "",
        }
    return {
        "product_name": product_name,
        "material": material,
        "category_hint": "",
        "keywords": suruga_keywords,
        "function": function_desc,
        "spec": spec,
    }


def _classify_one(img_file, text_ctx: str, ch_key: str) -> dict:
    """1件分の画像解析 + HTS照合を実行してresult dictを返す。"""
    image_bytes = img_file.getvalue()
    image_name  = img_file.name
    suruga_kw   = get_extra_keywords(text_ctx)

    # 品名に代表語があれば画像APIを呼ばずテキスト判定（コスト・レート節約）
    _term, _info = _decisive_match(text_ctx)
    if _term:
        return _classify_text_only(image_bytes, image_name, text_ctx, _term, _info, suruga_kw)

    if ch_key == AUTO_KEY:
        # 1回のAPIで「解析＋章推定」を同時取得（旧2回API→1回に統合）
        image_analysis, detected_all = analyze_and_predict(
            image_bytes, image_name, text_ctx, SUPPORTED_CHAPTERS
        )
        cache_hit_l1 = bool(image_analysis and image_analysis.get("_cache_hit"))
        query_list = [_build_query(image_analysis, suruga_kw)]

        # 画像ヒントに基づく章の除外フィルタを後段で適用
        # 日本語→英語の正規化も適用し、日本語カテゴリでも章ヒントが効くようにする
        img_hint_lower = _normalize_hint_app(" ".join(filter(None, [
            (image_analysis or {}).get("category_hint", ""),
            (image_analysis or {}).get("function", ""),
            " ".join((image_analysis or {}).get("keywords", [])),
        ])))
        exclude_chs: set[str] = set()
        for hint_key, excl_list in _HINT_EXCLUDE_CHAPTERS.items():
            if hint_key in img_hint_lower:
                exclude_chs.update(excl_list)

        # ヒントに応じて候補章を強制追加（誤った単一章のみへの誘導を防ぐ）
        include_chs: list[str] = []
        for hint_key, inc_list in _HINT_INCLUDE_CHAPTERS.items():
            if hint_key in img_hint_lower:
                include_chs.extend(inc_list)

        detected = [c for c in detected_all if c not in exclude_chs]
        for c in include_chs:
            if c not in detected and c not in exclude_chs and c in SUPPORTED_CHAPTERS:
                detected.append(c)
        detected = detected[:3]
        if not detected:
            # 除外で全滅した場合は元の推定を採用（安全側）
            detected = detected_all[:3]

        if not detected:
            # ── 最終フォールバック: 既ダウンロード済み全章でスキャン ──────────────────
            # AI の chapter_hints が非対応章のみ / surugaya DB が未知商品 / 翻訳失敗による
            # フィールド空白化が重なると detected_all = [] になりここに到達する。
            # 手動ホワイトリスト(_HINT_INCLUDE_CHAPTERS)の管理を不要にするため、
            # ローカルに存在する全章データを候補として使い、スコアリングで正しい章を選ぶ。
            _fallback_all = [
                k for k in SUPPORTED_CHAPTERS
                if (DATA_DIR / SUPPORTED_CHAPTERS[k]["data_file"]).exists()
            ]
            if not _fallback_all:
                return {"error": "章を推定できませんでした", "filename": image_name,
                        "image_bytes": image_bytes,
                        "image_analysis": image_analysis, "cache_hit_l1": cache_hit_l1}
            # 全章でスキャンし、スコア上位3章のみに絞って通常パスに流す
            _fallback_files = [(k, SUPPORTED_CHAPTERS[k]["data_file"]) for k in _fallback_all]
            _fallback_results = classify_per_chapter_ensemble(
                query_list, chapter_files=_fallback_files, top_n_per_chapter=1
            )
            # 各章のトップスコアで降順ソートし、上位3章を detected に採用
            _ch_scores = {
                ch: max((r.get("effective_score", 0) for r in rs), default=0)
                for ch, rs in _fallback_results.items() if rs
            }
            detected = sorted(_ch_scores, key=_ch_scores.get, reverse=True)[:3]
            if not detected:
                return {"error": "章を推定できませんでした", "filename": image_name,
                        "image_bytes": image_bytes,
                        "image_analysis": image_analysis, "cache_hit_l1": cache_hit_l1}

        # 未取得章をダウンロード
        for k in detected:
            if not (DATA_DIR / SUPPORTED_CHAPTERS[k]["data_file"]).exists():
                _download_chapter(k, SUPPORTED_CHAPTERS[k]["data_file"])

        chapter_files = [
            (k, SUPPORTED_CHAPTERS[k]["data_file"])
            for k in detected
            if (DATA_DIR / SUPPORTED_CHAPTERS[k]["data_file"]).exists()
        ]

        # L2キャッシュ
        cache_hit_l2 = False
        l2_cached = analysis_cache.get_cached_hts(image_analysis) if image_analysis else None
        if l2_cached:
            results = l2_cached
            cache_hit_l2 = True
        else:
            results = classify_per_chapter_ensemble(
                query_list, chapter_files=chapter_files, top_n_per_chapter=3
            )
            results = apply_hts_overrides(results, query_list)
            if image_analysis:
                analysis_cache.save_hts(image_analysis, results)

        return {
            "filename": image_name,
            "image_bytes": image_bytes,
            "image_analysis": image_analysis,
            "results": results,
            "detected_chapters": detected,
            "is_auto": True,
            "cache_hit_l1": cache_hit_l1,
            "cache_hit_l2": cache_hit_l2,
        }
    else:
        # 手動で章を選択 → 章推定は不要。画像解析のみ（API1回）
        analyses = analyze_image_ensemble(image_bytes, image_name, text_ctx, n=1)
        image_analysis = analyses[0] if analyses else None
        cache_hit_l1 = bool(image_analysis and image_analysis.get("_cache_hit"))
        query_list = [_build_query(image_analysis, suruga_kw)]

        cache_hit_l2 = False
        l2_cached = analysis_cache.get_cached_hts(image_analysis) if image_analysis else None
        if l2_cached:
            results = l2_cached
            cache_hit_l2 = True
        else:
            results = classify_ensemble(
                query_list, top_n=5,
                chapter_file=SUPPORTED_CHAPTERS[ch_key]["data_file"],
            )
            if image_analysis:
                analysis_cache.save_hts(image_analysis, results)

        return {
            "filename": image_name,
            "image_bytes": image_bytes,
            "image_analysis": image_analysis,
            "results": results,
            "detected_chapters": [],
            "is_auto": False,
            "cache_hit_l1": cache_hit_l1,
            "cache_hit_l2": cache_hit_l2,
        }


# ── 判定実行 ─────────────────────────────────────────────────────────────────
if st.button("判定する", type="primary"):
    if not uploaded_images and not text_context:
        st.warning("画像またはテキストを入力してください。")
    elif chapter_key == AUTO_KEY and not get_api_key() and not _MOCK_MODE:
        st.error("自動判定にはANTHROPIC_API_KEYが必要です。")
    else:
        n_items = len(uploaded_images) if uploaded_images else 1
        progress = st.progress(0, text=f"0 / {n_items} 件処理中...")

        batch_results = []

        if uploaded_images:
            completed = 0
            with ThreadPoolExecutor(max_workers=min(n_items, 2)) as executor:
                futures = {
                    executor.submit(_classify_one, img, text_context, chapter_key): img.name
                    for img in uploaded_images
                }
                for future in as_completed(futures):
                    try:
                        batch_results.append(future.result())
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        batch_results.append({"error": str(e), "filename": futures[future]})
                    completed += 1
                    progress.progress(completed / n_items, text=f"{completed} / {n_items} 件処理中...")
            # ファイル名順にソート
            order = {img.name: i for i, img in enumerate(uploaded_images)}
            batch_results.sort(key=lambda r: order.get(r.get("filename", ""), 999))
        else:
            # 画像なし・テキストのみ
            suruga_kw = get_extra_keywords(text_context)
            query_list = [_build_query(None, suruga_kw)]
            results = classify_ensemble(
                query_list, top_n=5,
                chapter_file=SUPPORTED_CHAPTERS.get(chapter_key, list(SUPPORTED_CHAPTERS.values())[0])["data_file"],
            ) if chapter_key != AUTO_KEY else []
            batch_results = [{"filename": "(テキストのみ)", "results": results,
                               "detected_chapters": [], "is_auto": False,
                               "cache_hit_l1": False, "cache_hit_l2": False}]
            progress.progress(1.0, text="完了")

        progress.empty()
        st.session_state["batch_results"] = batch_results

# ── 結果表示 ─────────────────────────────────────────────────────────────────
def _priority_mark(r: dict) -> str:
    ratio = r.get("own_match_ratio", 0.0)
    if ratio >= 0.5:
        return "◎"
    if ratio >= 0.2:
        return "〇"
    return "△"


def _build_excel(batch_results: list[dict]) -> bytes:
    """判定結果一覧をExcelファイルとして返す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HTS判定結果"

    # ── ヘッダー ──────────────────────────────────────────────
    headers = [
        "ファイル名", "採用HTSコード", "説明（英語）", "説明（日本語）",
        "スコア", "マッチ率", "CPSC", "メモ",
        "候補2位", "候補3位",
    ]
    header_fill = PatternFill("solid", fgColor="1A6EB5")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28

    # ── データ行 ──────────────────────────────────────────────
    fv = st.session_state.get("form_version", 0)
    alt_fill = PatternFill("solid", fgColor="EEF4FB")

    for row_idx, item in enumerate(batch_results, 2):
        filename = item.get("filename", "")
        if "error" in item:
            ws.cell(row=row_idx, column=1, value=filename)
            ws.cell(row=row_idx, column=2, value=f"エラー: {item['error']}")
            continue

        if item.get("is_auto") and isinstance(item.get("results"), dict):
            flat = [r for ch_r in item["results"].values() for r in ch_r]
        elif isinstance(item.get("results"), list):
            flat = item["results"]
        else:
            flat = []

        if not flat:
            ws.cell(row=row_idx, column=1, value=filename)
            ws.cell(row=row_idx, column=2, value="候補なし")
            continue

        safe_name = filename.replace(".", "_").replace(" ", "_")
        chosen_key = f"chosen_{safe_name}_{fv}"
        note_key   = f"note_{safe_name}_{fv}"
        chosen_code = st.session_state.get(chosen_key) or flat[0]["hts_code"]
        note        = st.session_state.get(note_key, "")

        chosen_entry = next((r for r in flat if r["hts_code"] == chosen_code), flat[0])
        lbl_raw = chosen_entry.get("description") or chosen_entry["full_description"].rsplit(" > ", 1)[-1]
        lbl_jp  = JP_LABELS.get(lbl_raw.rstrip(":").strip(), "")
        score   = chosen_entry.get("effective_score", chosen_entry.get("score", 0))
        ratio   = chosen_entry.get("own_match_ratio", 0)
        cpsc    = "⚠️ CPSC" if check_cpsc(chosen_code)["is_cpsc"] else ""

        others = [r for r in flat if r["hts_code"] != chosen_code][:2]

        row_data = [
            filename, chosen_code, lbl_raw.rstrip(":").strip(), lbl_jp,
            round(score, 3), round(ratio, 3), cpsc, note,
            others[0]["hts_code"] if len(others) > 0 else "",
            others[1]["hts_code"] if len(others) > 1 else "",
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (3, 4)))
            if fill:
                cell.fill = fill

    # ── 列幅 ──────────────────────────────────────────────────
    col_widths = [28, 18, 45, 35, 8, 8, 8, 20, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_result_card(r: dict, index: int):
    with st.container(border=True):
        own_label = r["description"] or r["full_description"].rsplit(" > ", 1)[-1]
        ancestor_breadcrumb = (
            r["full_description"].rsplit(" > ", 1)[0]
            if r["description"] and " > " in r["full_description"] else ""
        )
        jp_label = _entry_label_ja(r)
        mark = _priority_mark(r)

        cpsc = check_cpsc(r["hts_code"])
        cpsc_badge = f" 🚨 CPSC対象（{cpsc['category']}）" if cpsc["is_cpsc"] else ""

        if jp_label:
            st.markdown(f"**{mark} {index}. `{r['hts_code']}` — {jp_label}**{cpsc_badge}")
            st.caption(f"原文: {own_label}")
        else:
            st.markdown(f"**{mark} {index}. `{r['hts_code']}` — {own_label}**{cpsc_badge}")

        st.caption(f"一致率: {r['match_rate']:.0%}　自身一致率: {r['own_match_ratio']:.0%}　優先度: {mark}")
        if ancestor_breadcrumb:
            with st.expander("上位分類"):
                st.caption(ancestor_breadcrumb)

        rate_cols = st.columns(3)
        for col, (label, value) in zip(rate_cols, [
            ("一般税率", r["general_rate"]),
            ("特別税率", r["special_rate"]),
            ("Column 2", r["other_rate"]),
        ]):
            col.markdown(
                f"<div style='font-size:0.8rem;color:gray;'>{label}</div>"
                f"<div style='font-size:0.95rem;'>{value or '-'}</div>",
                unsafe_allow_html=True,
            )
        st.caption("マッチキーワード: " + ", ".join(r["matched_keywords"]))


batch_results = st.session_state.get("batch_results", [])
if batch_results:
    st.divider()
    _hdr_col, _dl_col = st.columns([3, 1])
    _hdr_col.subheader(f"判定結果（{len(batch_results)}件）")
    with _dl_col:
        _xlsx = _build_excel(batch_results)
        _ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        st.download_button(
            "📥 Excelダウンロード",
            data=_xlsx,
            file_name=f"hts_results_{_ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    for item in batch_results:
        filename = item.get("filename", "")
        with st.expander(f"📦 {filename}", expanded=True):
            # エラー
            if "error" in item:
                st.error(item["error"])
                if item.get("error") == "章を推定できませんでした":
                    st.info(
                        "対象22章の中に該当が見つかりませんでした。下の解析結果を参考に、"
                        "上の「対象Chapter」で章を手動選択して再判定してください。"
                    )
                ecol1, ecol2 = st.columns([1, 4])
                if item.get("image_bytes"):
                    ecol1.image(item["image_bytes"], width='stretch')
                if item.get("image_analysis"):
                    with ecol2.expander("画像解析結果（参考）", expanded=True):
                        st.json(_analysis_ja_view(item["image_analysis"]))
                continue

            results  = item.get("results")
            is_auto  = item.get("is_auto", False)
            detected = item.get("detected_chapters", [])

            # flat を先に組み立て
            if is_auto and isinstance(results, dict):
                flat = [r for ch_r in results.values() for r in ch_r]
            elif isinstance(results, list):
                flat = results
            else:
                flat = []

            # ── 採用HTSコード（上部・画像左横） ──────────────────────────
            if flat:
                safe_name = filename.replace(".", "_").replace(" ", "_")
                best_idx  = max(
                    range(len(flat)),
                    key=lambda i: flat[i].get("effective_score", flat[i].get("score", 0)),
                )

                def _opt_label(code: str, _flat=flat) -> str:
                    if code == "該当なし/その他":
                        return code
                    e = next((r for r in _flat if r["hts_code"] == code), None)
                    if not e:
                        return code
                    label = f"{code}　{_entry_label_ja(e)}"
                    if check_cpsc(code)["is_cpsc"]:
                        label = f"⚠️ {label}"
                    return label

                img_col, adopt_col = st.columns([1, 4])
                with img_col:
                    if item.get("image_bytes"):
                        st.image(item["image_bytes"], width='stretch')
                with adopt_col:
                    input_cols = st.columns([6, 2, 1])
                    chosen = input_cols[0].selectbox(
                        "採用HTSコード",
                        options=[r["hts_code"] for r in flat] + ["該当なし/その他"],
                        index=best_idx,
                        format_func=_opt_label,
                        key=f"chosen_{safe_name}_{st.session_state['form_version']}",
                    )
                    note = input_cols[1].text_input(
                        "メモ", key=f"note_{safe_name}_{st.session_state['form_version']}"
                    )
                    if input_cols[2].button(
                        "保存", key=f"save_{safe_name}_{st.session_state['form_version']}"
                    ):
                        file_exists = FEEDBACK_LOG.exists()
                        with open(FEEDBACK_LOG, "a", newline="", encoding="utf-8") as f:
                            writer = csv.writer(f)
                            if not file_exists:
                                writer.writerow(["timestamp", "filename", "chosen_code", "note"])
                            writer.writerow([
                                datetime.datetime.now().isoformat(), filename, chosen, note
                            ])
                        st.success("保存しました。")

                _chosen_entry = next((r for r in flat if r["hts_code"] == chosen), None)
                _jp = _entry_label_ja(_chosen_entry)
                if _jp:
                    st.caption(f"📋 {_jp}")

                chosen_cpsc = check_cpsc(chosen)
                if chosen_cpsc["is_cpsc"]:
                    st.warning(f"⚠️ 採用コード `{chosen}` はCPSC電子証明書（eFiling）の提出が必要な可能性があります。（カテゴリ: {chosen_cpsc['category']}）")

                if item.get("image_analysis"):
                    with st.expander("画像解析結果（参考）"):
                        st.json(_analysis_ja_view(item["image_analysis"]))

                st.divider()

            # ── キャッシュ・Chapter情報 ────────────────────────────────────
            if item.get("text_only"):
                st.info(f"📝 テキスト判定（品名「{item.get('decisive_term','')}」により画像解析をスキップ＝API課金なし）")
            elif item.get("cache_hit_l1"):
                st.info("⚡ L1キャッシュヒット（同一画像 — APIスキップ）")
            elif item.get("cache_hit_l2"):
                st.info("⚡ L2キャッシュヒット（同じ分析結果 — 照合スキップ）")

            if detected:
                st.caption("推定Chapter: " + "、".join(SUPPORTED_CHAPTERS[k]["label"] for k in detected))

            # ── HTS候補カード（下部） ──────────────────────────────────────
            if not flat:
                st.info("候補が見つかりませんでした。")
            elif is_auto and isinstance(results, dict):
                for ch_key_r in detected:
                    ch_results = results.get(ch_key_r, [])
                    if ch_results:
                        st.markdown(f"**{SUPPORTED_CHAPTERS[ch_key_r]['label']}**")
                        for i, r in enumerate(ch_results, 1):
                            _render_result_card(r, i)
            else:
                for i, r in enumerate(flat, 1):
                    _render_result_card(r, i)

